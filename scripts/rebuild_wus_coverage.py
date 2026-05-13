import json
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / 'WUS(金控)國泰人壽鑽美滿利率變動型美元終身壽險(定期給付型)-11501版(宣告4.25%-1150101啟用)(保費融資)_i值.xlsx'
OUT = ROOT / 'wus_coverage_data.js'

BASE_FACE = 12_000_000
FACE_UNITS = 12_000
UNIT_FACE = 1_000
PRE_RATE = 0.0175
DECLARED_RATE = 0.0425
POLICY_END_AGE = 104
ADJUST_AGE_LIMIT = 15
FORCE_ADD_YEAR_LIMIT = 6
BASE_FACE_AMOUNT = BASE_FACE


def xl_round(value, ndigits=0):
    # Excel ROUND: halves away from zero. Values here are positive.
    factor = 10 ** ndigits
    return int(value * factor + 0.5) / factor


def fmt_code(gender_code, issue_age, year):
    return f'{gender_code}WUS00{issue_age:02d}{year:03d}-001'


def c_factor(year):
    if year <= 2:
        return 0.0
    return xl_round(0.98 ** (year - 3), 4)


def threshold(attained_age):
    if attained_age <= 30:
        return 2.1
    if attained_age <= 40:
        return 1.8
    if attained_age <= 50:
        return 1.6
    if attained_age <= 60:
        return 1.3
    if attained_age <= 70:
        return 1.2
    if attained_age <= 90:
        return 1.05
    return 1.0


def load_tables():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ws501 = wb['501']
    ws517 = wb['517']
    wsgp = wb['GP']
    t501 = {}
    for row in ws501.iter_rows(values_only=True):
        key = row[0]
        if key:
            t501[str(key)] = row
    t517 = {}
    for row in ws517.iter_rows(values_only=True):
        key = row[0]
        if key:
            t517[str(key)] = row
    gp = {}
    for row in wsgp.iter_rows(values_only=True):
        key = row[0]
        if key:
            gp[str(key)] = row[9]
    return t501, t517, gp


def gp_rate(gender_code, issue_age, gp):
    key = f'{gender_code}WUS00{issue_age:02d}-015'
    if key not in gp:
        raise KeyError(f'Missing GP key {key}')
    return float(gp[key])


def lookup_501(t501, code, suffix, col_1_based):
    key = f'{code}{suffix}'
    if key not in t501:
        raise KeyError(f'Missing 501 key {key}')
    value = t501[key][col_1_based - 1]
    return float(value or 0)


def lookup_517(t517, code, col_1_based):
    if code not in t517:
        raise KeyError(f'Missing 517 key {code}')
    value = t517[code][col_1_based - 1]
    return float(value or 0) / 10000


def build_series(gender_code, issue_age, t501, t517, gp):
    rate = gp_rate(gender_code, issue_age, gp)
    basic_premium = rate * FACE_UNITS
    values = []
    x_total_add = 0.0
    s_child_accum = 0.0
    ai_basic_survival = 0.0
    # Cache per-year 517 initial/terminal and 501 values for next-year lookups.
    rows = {}
    max_year = POLICY_END_AGE - issue_age + 1
    for year in range(1, max_year + 2):
        attained = issue_age + year - 1
        code = fmt_code(gender_code, issue_age, year)
        try:
            rows[year] = {
                'attained': attained,
                'g': lookup_501(t501, code, 2, 11),
                'h': lookup_501(t501, code, 2, 12),
                'i': lookup_501(t501, code, 1, 11),
                'j': lookup_501(t501, code, 1, 12),
                'l': lookup_517(t517, code, 9),
                'm': lookup_517(t517, code, 10),
            }
        except KeyError:
            if year <= max_year:
                raise
            rows[year] = None

    for year in range(1, max_year + 1):
        row = rows[year]
        next_row = rows.get(year + 1) or row
        attained = row['attained']
        cf = c_factor(year)
        e_factor = 1.02
        bw = threshold(attained)
        x_prev = x_total_add

        # Dividend sharing. WUS uses terminal value basis in this workbook.
        p_share = (DECLARED_RATE - PRE_RATE) * (row['j'] * FACE_UNITS + x_prev * row['m'])
        r_child = p_share if attained < ADJUST_AGE_LIMIT else 0.0
        if attained <= ADJUST_AGE_LIMIT:
            s_child_accum = s_child_accum * ((1 + DECLARED_RATE / 12) ** 12) + r_child
        else:
            s_child_accum = 0.0
        t_child_add = (s_child_accum / next_row['l']) if attained == ADJUST_AGE_LIMIT and next_row['l'] else 0.0

        current_add = 0.0
        if attained >= ADJUST_AGE_LIMIT and next_row['l']:
            current_add += p_share / next_row['l']
        current_add += t_child_add
        current_add = xl_round(current_add, 2)
        x_total_add = x_prev + current_add

        # Basic and added-account survival accumulation.
        ah_basic_survival = xl_round((row['j'] - next_row['i']) * FACE_UNITS, 0)
        ai_basic_survival = ah_basic_survival if year == 1 else ai_basic_survival + ah_basic_survival

        ab_basic_face = xl_round(BASE_FACE * cf, 0)
        ac_added_face = xl_round(cf * x_prev, 0)  # workbook setting: do not include current-year small add.

        ae_basic_cash = xl_round(xl_round(row['j'] * FACE_UNITS, 0) * bw, 0)
        af_added_cash = xl_round(xl_round(x_prev * next_row['l'], 0) * bw, 0)

        aj_premium_basis = xl_round(basic_premium, 0) * min(year, 1) * e_factor - ai_basic_survival
        ak_added_premium_basis = xl_round(x_prev * rate / UNIT_FACE, 0) * min(year, 1) * e_factor - (x_prev * ai_basic_survival / BASE_FACE)

        premium_basis = xl_round(aj_premium_basis + ak_added_premium_basis, 2)
        cash_basis = ae_basic_cash + af_added_cash
        face_basis = ab_basic_face + ac_added_face
        total_death = max(premium_basis, cash_basis, face_basis)
        values.append(int(xl_round(total_death, 0)))
    return values


def main():
    t501, t517, gp = load_tables()
    data = {'male': {}, 'female': {}}
    for gender_key, gender_code in [('male', 1), ('female', 2)]:
        for age in range(0, 76):
            data[gender_key][str(age)] = build_series(gender_code, age, t501, t517, gp)
    data['dividendOption'] = '第7年起持續增購保額'
    data['baseFaceAmount'] = BASE_FACE_AMOUNT
    data['unitFaceAmount'] = UNIT_FACE
    data['source'] = 'Rebuilt from WUS workbook 501/517/GP tables using AP total death benefit logic and add_method=1 增額繳清'
    text = (
        '// Generated from WUS 鑽美滿 workbook using the 第7年起持續增購保額 dividend option.\n'
        '// Values are total death/disability benefit by issue age, gender, and policy year.\n'
        'window.WUS_COVERAGE_BY_AGE_GENDER = '
        + json.dumps(data, ensure_ascii=False, separators=(',', ':'))
        + ';\n'
    )
    OUT.write_text(text, encoding='utf-8')

if __name__ == '__main__':
    main()

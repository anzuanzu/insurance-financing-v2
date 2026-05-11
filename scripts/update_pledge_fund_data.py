#!/usr/bin/env python3
"""Generate pledge_fund_data.js from a pledgeable fund xlsx file.

Expected xlsx format:
- Row 1 is headers.
- Column A contains fund product codes.
- Column C contains fund Chinese names.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def normalize_code(value) -> str:
    return str(value or "").strip().upper()


def normalize_name(value) -> str:
    return str(value or "").strip()


def read_pledge_funds(xlsx_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    products: list[dict[str, str]] = []
    seen: set[str] = set()

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        code = normalize_code(row[0] if len(row) > 0 else "")
        name = normalize_name(row[2] if len(row) > 2 else "")
        if not code or not name or code in seen:
            continue
        seen.add(code)
        products.append({"code": code, "name": name})

    return products


def write_js(products: list[dict[str, str]], output_path: Path, source_name: str) -> None:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    json_text = json.dumps(products, ensure_ascii=False, indent=2)
    content = (
        "// pledge_fund_data.js - 最新可質借基金清單\n"
        "// 由 scripts/update_pledge_fund_data.py 依 xlsx A 欄「商品代碼」與 C 欄「基金中文名稱」產生。\n"
        f"// Source: {source_name}\n"
        f"// Generated at: {generated_at}\n"
        f"const pledgeFundProducts = {json_text};\n"
    )
    output_path.write_text(content, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate pledge_fund_data.js from xlsx.")
    parser.add_argument("xlsx", type=Path, help="Path to latest pledgeable fund xlsx file")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("pledge_fund_data.js"),
        help="Output JS file path",
    )
    args = parser.parse_args()

    products = read_pledge_funds(args.xlsx)
    if not products:
        raise SystemExit("No fund products found. Check that column A has codes and column C has Chinese names.")

    write_js(products, args.output, args.xlsx.name)
    print(f"Wrote {len(products)} fund products to {args.output}")


if __name__ == "__main__":
    main()

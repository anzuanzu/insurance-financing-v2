import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class WorkbookCellMap:
    gender: str
    age: str
    face_amount: str
    dividend_option: str


@dataclass(frozen=True)
class WorkbookRangeMap:
    paid_premium: str
    table_premium: str
    total_benefit: str


def normalize_face_amount_to_thousand_usd(face_amount: int) -> int:
    return int(round(face_amount / 1000))


def write_inputs(workbook_path: Path, sheet_name: str, cell_map: WorkbookCellMap, *, gender_code: int, age: int, face_amount: int, dividend_option_code: int) -> None:
    workbook = load_workbook(workbook_path, data_only=False)
    sheet = workbook[sheet_name]
    sheet[cell_map.gender] = gender_code
    sheet[cell_map.age] = age
    sheet[cell_map.face_amount] = normalize_face_amount_to_thousand_usd(face_amount)
    sheet[cell_map.dividend_option] = dividend_option_code
    workbook.save(workbook_path)
    workbook.close()


def recalc_with_libreoffice(workbook_path: Path) -> Path:
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise RuntimeError("LibreOffice is not available.")

    out_dir = workbook_path.parent / "recalc"
    out_dir.mkdir(parents=True, exist_ok=True)

    command = [
        soffice,
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(out_dir),
        str(workbook_path),
    ]
    subprocess.run(command, check=True, capture_output=True, text=True)
    recalculated_path = out_dir / workbook_path.name
    if not recalculated_path.exists():
        raise RuntimeError("LibreOffice did not produce a recalculated workbook.")
    return recalculated_path


def read_scalar(workbook_path: Path, ref: str) -> float:
    sheet_name, cell = ref.split("!")
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    value = workbook[sheet_name][cell].value
    workbook.close()
    return float(value or 0)


def read_range_value(workbook_path: Path, ref: str, index: int) -> float:
    sheet_name, cell_range = ref.split("!")
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook[sheet_name]
    cells = list(sheet[cell_range])
    flat = [cell.value for row in cells for cell in row]
    workbook.close()
    if index < 0 or index >= len(flat):
        raise IndexError(f"Range index {index} out of bounds for {ref}.")
    return float(flat[index] or 0)


def create_temp_workbook_copy(source_workbook: Path) -> Path:
    temp_dir = Path(tempfile.mkdtemp(prefix="wus-workbook-"))
    target = temp_dir / source_workbook.name
    shutil.copy2(source_workbook, target)
    return target
